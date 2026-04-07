from openpyxl import load_workbook

def load_career_info(path=r'..\data\Career Dataset.xlsx'):
    """
    Returns a dict like:
    {
      'Software Development and Engineering': ['Web Development', 'AI', ...],
      'Data Science': [...],
      ...
    }
    """
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    career_map = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:  # skip header
        if row[0] and row[1]:
            skills = [s.strip() for s in row[1].split(',')]
            career_map[row[0]] = skills
    return career_map

# Maps specific job roles to broad career categories
ROLE_TO_CAREER = {
    'Database Administrator':        'Software Development and Engineering',
    'Hardware Engineer':             'Software Development and Engineering',
    'Application Support Engineer':  'Software Development and Engineering',
    'Cyber Security Specialist':     'Security',
    'Information Security Specialist':'Security',
    'Networking Engineer':           'Software Development and Engineering',
    'Software Developer':            'Software Development and Engineering',
    'API Specialist':                'Software Development and Engineering',
    'Project Manager':               'Development',
    'Technical Writer':              'User Experience (UX) and User Interface (UI) Design',
    'AI ML Specialist':              'Artificial Intelligence',
    'Software tester':               'Software Development and Engineering',
    'Business Analyst':              'Development',
    'Customer Service Executive':    'Development',
    'Data Scientist':                'Data Science',
    'Helpdesk Engineer':             'Software Development and Engineering',
    'Graphics Designer':             'User Experience (UX) and User Interface (UI) Design',
}

def get_career_path(role, career_map):
    category = ROLE_TO_CAREER.get(role, None)
    if not category or category not in career_map:
        return None, []
    return category, career_map[category]